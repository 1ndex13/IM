import io
import openpyxl
from datetime import datetime, timedelta
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count, F, Q
from django.core.paginator import Paginator
from products.models import Product, StockTransaction, Category
from expenses.models import ExpenseInvoiceItem


@login_required
def stock_report(request):
    category_filter = request.GET.get('category', '')
    low_stock_only = request.GET.get('low_stock', '')
    export = request.GET.get('export', '')

    products = Product.objects.all().select_related('category')

    if category_filter:
        products = products.filter(category_id=category_filter)

    if low_stock_only:
        products = products.filter(quantity__lte=F('min_stock'))

    categories = Category.objects.all()

    if export == 'excel':
        return export_stock_excel(products)

    total_products = products.count()
    low_stock_count = products.filter(quantity__lte=F('min_stock')).count()
    zero_stock_count = products.filter(quantity=0).count()
    normal_stock_count = total_products - low_stock_count

    paginator = Paginator(products, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'products': page_obj,
        'categories': categories,
        'category_filter': category_filter,
        'low_stock_only': low_stock_only,
        'total_products': total_products,
        'low_stock_count': low_stock_count,
        'zero_stock_count': zero_stock_count,
        'normal_stock_count': normal_stock_count,
    }
    return render(request, 'reports/stock_report.html', context)


@login_required
def movement_report(request):
    product_id = request.GET.get('product', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    transaction_type = request.GET.get('type', '')
    export = request.GET.get('export', '')

    transactions = StockTransaction.objects.all().select_related('product', 'user')

    if product_id:
        transactions = transactions.filter(product_id=product_id)

    if date_from:
        transactions = transactions.filter(date__gte=date_from)

    if date_to:
        transactions = transactions.filter(date__lte=date_to)

    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)

    products = Product.objects.all()

    if export == 'excel':
        return export_movement_excel(transactions)

    total_transactions = transactions.count()
    in_count = transactions.filter(transaction_type='in').count()
    out_count = transactions.filter(transaction_type='out').count()
    unique_products = transactions.values('product').distinct().count()

    paginator = Paginator(transactions, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'transactions': page_obj,
        'products': products,
        'product_id': product_id,
        'date_from': date_from,
        'date_to': date_to,
        'transaction_type': transaction_type,
        'total_transactions': total_transactions,
        'in_count': in_count,
        'out_count': out_count,
        'unique_products': unique_products,
    }
    return render(request, 'reports/movement_report.html', context)


@login_required
def turnover_report(request):
    period = request.GET.get('period', '30')
    category_filter = request.GET.get('category', '')
    export = request.GET.get('export', '')

    end_date = datetime.now()
    start_date = end_date - timedelta(days=int(period))

    incoming_data = StockTransaction.objects.filter(
        transaction_type='in',
        date__range=[start_date, end_date]
    ).values('product__name', 'product__sku').annotate(
        total_quantity=Sum('quantity'),
        transaction_count=Count('id')
    ).order_by('-total_quantity')

    outgoing_data = StockTransaction.objects.filter(
        transaction_type='out',
        date__range=[start_date, end_date]
    ).values('product__name', 'product__sku').annotate(
        total_quantity=Sum('quantity'),
        transaction_count=Count('id')
    ).order_by('-total_quantity')

    popular_products = list(outgoing_data[:10])

    all_products = Product.objects.all()
    if category_filter:
        all_products = all_products.filter(category_id=category_filter)

    active_product_ids = StockTransaction.objects.filter(
        date__range=[start_date, end_date]
    ).values_list('product_id', flat=True).distinct()

    inactive_products = all_products.exclude(id__in=active_product_ids)

    categories = Category.objects.all()

    if export == 'excel':
        return export_turnover_excel(popular_products, inactive_products, period)

    context = {
        'popular_products': popular_products,
        'inactive_products': inactive_products,
        'period': period,
        'categories': categories,
        'category_filter': category_filter,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'reports/turnover_report.html', context)


def export_stock_excel(products):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Остатки товаров"

    headers = ['Артикул', 'Наименование', 'Категория', 'Ед. изм.', 'Цена', 'Текущий остаток', 'Мин. запас', 'Статус']
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)

    for row, product in enumerate(products, 2):
        status = "Низкий запас" if product.quantity <= product.min_stock else "Норма"

        worksheet.cell(row=row, column=1, value=product.sku)
        worksheet.cell(row=row, column=2, value=product.name)
        worksheet.cell(row=row, column=3, value=str(product.category) if product.category else '')
        worksheet.cell(row=row, column=4, value=product.get_unit_display())
        worksheet.cell(row=row, column=5, value=float(product.price))
        worksheet.cell(row=row, column=6, value=product.quantity)
        worksheet.cell(row=row, column=7, value=product.min_stock)
        worksheet.cell(row=row, column=8, value=status)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="stock_report.xlsx"'

    workbook.save(response)
    return response


def export_movement_excel(transactions):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Движение товаров"

    headers = ['Дата', 'Товар', 'Артикул', 'Тип операции', 'Количество', 'Пользователь', 'Комментарий']
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)

    for row, transaction in enumerate(transactions, 2):
        worksheet.cell(row=row, column=1, value=transaction.date.strftime("%d.%m.%Y %H:%M"))
        worksheet.cell(row=row, column=2, value=transaction.product.name)
        worksheet.cell(row=row, column=3, value=transaction.product.sku)
        worksheet.cell(row=row, column=4, value=transaction.get_transaction_type_display())
        worksheet.cell(row=row, column=5, value=transaction.quantity)
        worksheet.cell(row=row, column=6, value=str(transaction.user))
        worksheet.cell(row=row, column=7, value=transaction.comment)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="movement_report.xlsx"'

    workbook.save(response)
    return response


def export_turnover_excel(popular_products, inactive_products, period):
    workbook = openpyxl.Workbook()

    ws1 = workbook.active
    ws1.title = "Популярные товары"

    headers = ['Товар', 'Артикул', 'Количество продаж', 'Количество операций']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=header)

    for row, product in enumerate(popular_products, 2):
        ws1.cell(row=row, column=1, value=product['product__name'])
        ws1.cell(row=row, column=2, value=product['product__sku'])
        ws1.cell(row=row, column=3, value=product['total_quantity'])
        ws1.cell(row=row, column=4, value=product['transaction_count'])

    ws2 = workbook.create_sheet("Товары без движения")

    headers = ['Товар', 'Артикул', 'Категория', 'Текущий остаток']
    for col, header in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=header)

    for row, product in enumerate(inactive_products, 2):
        ws2.cell(row=row, column=1, value=product.name)
        ws2.cell(row=row, column=2, value=product.sku)
        ws2.cell(row=row, column=3, value=str(product.category) if product.category else '')
        ws2.cell(row=row, column=4, value=product.quantity)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="turnover_report_{period}d.xlsx"'

    workbook.save(response)
    return response